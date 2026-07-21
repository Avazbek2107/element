from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
from datetime import date as date_type, timedelta
import csv, io, secrets
from app.database import get_db
from app.models.user import User, UserRole
from app.models.student import StudentProfile
from app.models.group import Group
from app.models.test import TestResult, TestStatus
from app.models.attendance import Attendance, AttendanceStatus
from app.schemas.student import StudentCreate, StudentUpdate, StudentOut, StudentListOut
from app.utils.auth import get_current_user, require_roles
from app.utils.auth import hash_password
from app.utils.audit import log_action

router = APIRouter(prefix="/api/students", tags=["students"])

AdminOrTeacher = require_roles(UserRole.admin, UserRole.teacher, module="students")

GRADES = [(90, "A'lo"), (75, "Yaxshi"), (50, "O'rtacha"), (0, "Yomon")]


def _grade(pct: float) -> str:
    for threshold, label in GRADES:
        if pct >= threshold:
            return label
    return "Yomon"


def _build_student_out(profile: StudentProfile) -> StudentOut:
    user = profile.user
    return StudentOut(
        id=profile.id,
        user_id=profile.user_id,
        first_name=user.first_name,
        last_name=user.last_name,
        middle_name=profile.middle_name,
        gender=profile.gender,
        email=user.email,
        phone=user.phone,
        group_id=profile.group_id,
        group_name=profile.group.name if profile.group else None,
        birth_date=profile.birth_date,
        parent_phone=profile.parent_phone,
        address=profile.address,
        avatar_url=profile.avatar_url,
        doc_type=profile.doc_type,
        doc_series=profile.doc_series,
        link_code=profile.link_code,
        student_link_code=profile.student_link_code,
        parent_telegram_id=profile.parent_telegram_id,
        student_telegram_id=profile.student_telegram_id,
        enrolled_date=profile.enrolled_date,
        course_start_date=profile.course_start_date,
        course_end_date=profile.course_end_date,
        is_active=user.is_active,
    )


@router.get("", response_model=StudentListOut)
def list_students(
    search: Optional[str] = Query(None),
    group_id: Optional[int] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: User = Depends(AdminOrTeacher),
):
    query = db.query(StudentProfile).join(User).filter(User.is_active == True)

    if group_id:
        query = query.filter(StudentProfile.group_id == group_id)

    if search:
        query = query.filter(
            (User.first_name.ilike(f"%{search}%"))
            | (User.last_name.ilike(f"%{search}%"))
            | (User.phone.ilike(f"%{search}%"))
        )

    total = query.count()
    profiles = query.offset(skip).limit(limit).all()
    return {"items": [_build_student_out(p) for p in profiles], "total": total}


@router.post("", response_model=StudentOut, status_code=status.HTTP_201_CREATED)
def create_student(
    body: StudentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(AdminOrTeacher),
):
    if db.query(User).filter(User.email == body.email).first():
        raise HTTPException(status_code=400, detail="Bu email allaqachon mavjud")
    if db.query(User).filter(User.username == body.username).first():
        raise HTTPException(status_code=400, detail="Bu username band")

    user = User(
        username=body.username,
        email=body.email,
        password_hash=hash_password(body.password),
        first_name=body.first_name,
        last_name=body.last_name,
        phone=body.phone,
        role=UserRole.student,
    )
    db.add(user)
    db.flush()

    profile = StudentProfile(
        user_id=user.id,
        group_id=body.group_id,
        middle_name=body.middle_name,
        gender=body.gender,
        birth_date=body.birth_date,
        parent_phone=body.parent_phone,
        parent_telegram_id=body.parent_telegram_id,
        address=body.address,
        doc_type=body.doc_type,
        doc_series=body.doc_series,
        course_start_date=body.course_start_date,
        course_end_date=body.course_end_date,
    )
    db.add(profile)
    db.flush()
    log_action(db, current_user, "create", "students", "student", profile.id, f"{user.first_name} {user.last_name}")
    db.commit()
    db.refresh(profile)
    return _build_student_out(profile)


@router.get("/{student_id}", response_model=StudentOut)
def get_student(
    student_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    profile = db.query(StudentProfile).filter(StudentProfile.id == student_id).first()
    if not profile:
        raise HTTPException(status_code=404, detail="O'quvchi topilmadi")
    # O'quvchi faqat o'zini ko'ra oladi
    if current_user.role == UserRole.student:
        own_profile = db.query(StudentProfile).filter(StudentProfile.user_id == current_user.id).first()
        if not own_profile or own_profile.id != student_id:
            raise HTTPException(status_code=403, detail="Ruxsat yo'q")
    return _build_student_out(profile)


@router.get("/{student_id}/progress")
def get_student_progress(
    student_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    profile = db.query(StudentProfile).filter(StudentProfile.id == student_id).first()
    if not profile:
        raise HTTPException(status_code=404, detail="O'quvchi topilmadi")
    if current_user.role == UserRole.student:
        own_profile = db.query(StudentProfile).filter(StudentProfile.user_id == current_user.id).first()
        if not own_profile or own_profile.id != student_id:
            raise HTTPException(status_code=403, detail="Ruxsat yo'q")

    # ── Test natijalari tarixi ────────────────────────────────
    results = (
        db.query(TestResult)
        .filter(TestResult.student_id == student_id, TestResult.status == TestStatus.submitted)
        .order_by(TestResult.submitted_at.asc())
        .all()
    )
    test_history = [
        {
            "test_id":      r.test_id,
            "title":        r.test.title if r.test else "",
            "test_type":    r.test.test_type.value if r.test and r.test.test_type else None,
            "percentage":   float(r.percentage),
            "grade":        _grade(float(r.percentage)),
            "submitted_at": r.submitted_at,
        }
        for r in results
    ]
    score_avg = round(sum(t["percentage"] for t in test_history) / len(test_history), 1) if test_history else None

    # ── Yo'qlama (so'nggi 6 oy, oylik) ────────────────────────
    today = date_type.today()
    attendance_monthly = []
    for i in range(5, -1, -1):
        m_index = today.month - 1 - i
        y = today.year + (m_index // 12)
        m = (m_index % 12) + 1
        att_m = db.query(Attendance).filter(
            Attendance.student_id == student_id,
            func.extract('year', Attendance.date) == y,
            func.extract('month', Attendance.date) == m,
        )
        present = att_m.filter(Attendance.status == AttendanceStatus.present).count()
        late    = att_m.filter(Attendance.status == AttendanceStatus.late).count()
        total   = att_m.count()
        attendance_monthly.append({
            "month": f"{y}-{m:02d}",
            "rate":  round((present + late) / total * 100, 1) if total else None,
        })

    overall_att = db.query(Attendance).filter(Attendance.student_id == student_id)
    overall_present = overall_att.filter(Attendance.status == AttendanceStatus.present).count()
    overall_late    = overall_att.filter(Attendance.status == AttendanceStatus.late).count()
    overall_total   = overall_att.count()
    attendance_rate_overall = round((overall_present + overall_late) / overall_total * 100, 1) if overall_total else None

    return {
        "test_history":            test_history,
        "score_avg":                score_avg,
        "attendance_monthly":       attendance_monthly,
        "attendance_rate_overall":  attendance_rate_overall,
    }


@router.put("/{student_id}", response_model=StudentOut)
def update_student(
    student_id: int,
    body: StudentUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(AdminOrTeacher),
):
    profile = db.query(StudentProfile).filter(StudentProfile.id == student_id).first()
    if not profile:
        raise HTTPException(status_code=404, detail="O'quvchi topilmadi")

    user = profile.user
    USER_FIELDS = {"first_name", "last_name", "phone"}
    update_data = body.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        if field in USER_FIELDS:
            setattr(user, field, value)
        else:
            setattr(profile, field, value)

    log_action(db, current_user, "update", "students", "student", profile.id, f"{user.first_name} {user.last_name}")
    db.commit()
    db.refresh(profile)
    return _build_student_out(profile)


@router.delete("/{student_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_student(
    student_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.admin, module="students")),
):
    profile = db.query(StudentProfile).filter(StudentProfile.id == student_id).first()
    if not profile:
        raise HTTPException(status_code=404, detail="O'quvchi topilmadi")
    label = f"{profile.user.first_name} {profile.user.last_name}" if profile.user else None
    log_action(db, current_user, "delete", "students", "student", profile.id, label)
    profile.user.is_active = False
    db.commit()


@router.post("/import", status_code=200)
async def import_students(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(AdminOrTeacher),
):
    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="Faqat .csv yoki .xlsx fayl yuklang")

    content = await file.read()

    if ext == "csv":
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = [{k.strip(): (v.strip() if v else "") for k, v in r.items()} for r in reader]
    else:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        label_map = {
            "Ism": "first_name", "Familiya": "last_name", "Sharif": "middle_name",
            "Jinsi": "gender", "Email": "email", "Telefon": "phone",
            "Login": "username", "Parol": "password",
            "Tug'ilgan sana": "birth_date", "Ota-ona tel": "parent_phone", "Manzil": "address",
        }
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        keys = [label_map.get(h, h.lower()) for h in headers]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rows.append({keys[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})

    created, skipped, errors = [], [], []

    for i, row in enumerate(rows, start=1):
        fn    = row.get("first_name", "").strip()
        ln    = row.get("last_name", "").strip()
        mn    = row.get("middle_name", "").strip() or None
        _g    = row.get("gender", "").strip().lower()
        GENDER_MAP = {"male": "male", "female": "female", "o'g'il": "male", "qiz": "female", "erkak": "male", "ayol": "female"}
        gender = GENDER_MAP.get(_g) or None
        email = row.get("email", "").strip()
        uname = row.get("username", "").strip()
        pwd   = row.get("password", "").strip() or secrets.token_urlsafe(12)
        phone = row.get("phone", "").strip() or None
        parent_phone = row.get("parent_phone", "").strip() or None
        address = row.get("address", "").strip() or None

        from datetime import date as dt_date
        birth_date = None
        bd_str = row.get("birth_date", "").strip()
        if bd_str:
            for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
                try:
                    from datetime import datetime
                    birth_date = datetime.strptime(bd_str, fmt).date()
                    break
                except Exception:
                    pass

        if not fn or not ln or not email or not uname:
            errors.append({"row": i, "reason": "Majburiy maydonlar bo'sh"})
            continue
        if db.query(User).filter(User.email == email).first():
            skipped.append({"row": i, "email": email, "reason": "Email allaqachon mavjud"})
            continue
        if db.query(User).filter(User.username == uname).first():
            skipped.append({"row": i, "email": email, "reason": "Username band"})
            continue

        user = User(
            first_name=fn, last_name=ln, email=email, username=uname,
            phone=phone, password_hash=hash_password(pwd), role=UserRole.student,
        )
        db.add(user)
        db.flush()

        profile = StudentProfile(
            user_id=user.id, middle_name=mn, gender=gender,
            birth_date=birth_date, parent_phone=parent_phone, address=address,
        )
        db.add(profile)
        created.append({"email": email, "username": uname})

    db.commit()
    return {"created": len(created), "skipped": len(skipped), "errors": len(errors),
            "details": {"skipped": skipped, "errors": errors}}
