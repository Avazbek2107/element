from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from pydantic import BaseModel
import csv, io
from pathlib import Path
from app.database import get_db
from app.models.user import User, UserRole
from app.utils.auth import require_roles, hash_password

router = APIRouter(prefix="/api/users", tags=["users"])

AdminOnly = require_roles(UserRole.admin, module="teachers")


class TeacherCreate(BaseModel):
    first_name: str
    last_name: str
    email: str
    username: str
    password: str
    phone: Optional[str] = None


class TeacherUpdate(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: Optional[str] = None
    username: Optional[str] = None
    password: Optional[str] = None
    phone: Optional[str] = None


def _parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        rows.append({k.strip(): v.strip() for k, v in row.items()})
    return rows


def _parse_xlsx(content: bytes) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [str(cell.value).strip() for cell in ws[1]]
    # Excel'dagi uzbek sarlavhalarni ingliz kalitlarga map qilamiz
    label_map = {
        "Ism": "first_name", "Familiya": "last_name", "Email": "email",
        "Telefon": "phone", "Login": "username", "Parol": "password",
        "Fan": "subject",
    }
    keys = [label_map.get(h, h.lower()) for h in headers]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rows.append({keys[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
    return rows


@router.post("/import-teachers")
async def import_teachers(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(AdminOnly),
):
    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="Faqat .csv yoki .xlsx fayl yuklang")

    content = await file.read()
    try:
        rows = _parse_csv(content) if ext == "csv" else _parse_xlsx(content)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Faylni o'qishda xato: {e}")

    if not rows:
        raise HTTPException(status_code=422, detail="Fayl bo'sh yoki format noto'g'ri")

    created, skipped, errors = [], [], []

    for i, row in enumerate(rows, start=1):
        fn   = row.get("first_name", "").strip()
        ln   = row.get("last_name", "").strip()
        email= row.get("email", "").strip()
        uname= row.get("username", "").strip()
        pwd  = row.get("password", "Teacher@1234")
        phone= row.get("phone", "").strip() or None

        if not fn or not ln or not email or not uname:
            errors.append({"row": i, "reason": "Majburiy maydonlar bo'sh (ism, familiya, email, username)"})
            continue

        if db.query(User).filter(User.email == email).first():
            skipped.append({"row": i, "email": email, "reason": "Email allaqachon mavjud"})
            continue
        if db.query(User).filter(User.username == uname).first():
            skipped.append({"row": i, "email": email, "reason": "Username band"})
            continue

        user = User(
            first_name=fn,
            last_name=ln,
            email=email,
            username=uname,
            phone=phone,
            password_hash=hash_password(pwd),
            role=UserRole.teacher,
        )
        db.add(user)
        created.append({"email": email, "username": uname})

    db.commit()

    return {
        "created": len(created),
        "skipped": len(skipped),
        "errors": len(errors),
        "details": {"created": created, "skipped": skipped, "errors": errors},
    }


@router.post("/teachers")
def create_teacher(
    body: TeacherCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(AdminOnly),
):
    if db.query(User).filter(User.email == body.email).first():
        raise HTTPException(status_code=400, detail="Bu email allaqachon mavjud")
    if db.query(User).filter(User.username == body.username).first():
        raise HTTPException(status_code=400, detail="Bu username band")
    if body.phone and db.query(User).filter(User.phone == body.phone).first():
        raise HTTPException(status_code=400, detail="Bu telefon raqami band")

    teacher = User(
        first_name=body.first_name,
        last_name=body.last_name,
        email=body.email,
        username=body.username,
        phone=body.phone or None,
        password_hash=hash_password(body.password),
        role=UserRole.teacher,
    )
    db.add(teacher)
    db.commit()
    db.refresh(teacher)
    return {"id": teacher.id, "first_name": teacher.first_name, "last_name": teacher.last_name,
            "email": teacher.email, "phone": teacher.phone, "username": teacher.username}


@router.put("/teachers/{teacher_id}")
def update_teacher(
    teacher_id: int,
    body: TeacherUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(AdminOnly),
):
    teacher = db.query(User).filter(User.id == teacher_id, User.role == UserRole.teacher).first()
    if not teacher:
        raise HTTPException(status_code=404, detail="O'qituvchi topilmadi")

    if body.email and body.email != teacher.email:
        if db.query(User).filter(User.email == body.email).first():
            raise HTTPException(status_code=400, detail="Bu email allaqachon mavjud")
        teacher.email = body.email

    if body.username and body.username != teacher.username:
        if db.query(User).filter(User.username == body.username).first():
            raise HTTPException(status_code=400, detail="Bu username band")
        teacher.username = body.username

    if body.phone is not None:
        if body.phone and body.phone != teacher.phone:
            if db.query(User).filter(User.phone == body.phone).first():
                raise HTTPException(status_code=400, detail="Bu telefon raqami band")
        teacher.phone = body.phone or None

    if body.first_name:
        teacher.first_name = body.first_name
    if body.last_name:
        teacher.last_name = body.last_name
    if body.password:
        teacher.password_hash = hash_password(body.password)

    db.commit()
    db.refresh(teacher)
    return {"id": teacher.id, "first_name": teacher.first_name, "last_name": teacher.last_name,
            "email": teacher.email, "phone": teacher.phone, "username": teacher.username}


@router.delete("/teachers/{teacher_id}")
def delete_teacher(
    teacher_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(AdminOnly),
):
    teacher = db.query(User).filter(User.id == teacher_id, User.role == UserRole.teacher).first()
    if not teacher:
        raise HTTPException(status_code=404, detail="O'qituvchi topilmadi")
    db.delete(teacher)
    db.commit()
    return {"ok": True}


@router.get("/teachers", response_model=List[dict])
def list_teachers(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.super_admin, UserRole.admin, UserRole.teacher)),
):
    teachers = db.query(User).filter(
        User.role == UserRole.teacher,
        User.is_active == True,
    ).order_by(User.last_name).all()
    return [
        {"id": t.id, "first_name": t.first_name, "last_name": t.last_name,
         "email": t.email, "phone": t.phone, "username": t.username}
        for t in teachers
    ]
